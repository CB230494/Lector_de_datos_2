# app.py
import streamlit as st
import pandas as pd
import sqlite3
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any

st.set_page_config(page_title="Avances por meta", layout="wide")
st.subheader("📈 Avances por meta")

# =========================
# 1) CONFIG DB & DATOS BASE
# =========================
DB_PATH = "avances.db"

# === PLAN BASE (de tu matriz) ===
# Mapeo:
# - actividad_estrategica -> actividad (para mostrar)
# - meta_cuantitativa     -> meta_total (para cálculos)
PLAN_BASE = [
    {
        "fila": 1,
        "indole": "Operativo",
        "actividad_estrategica": "Coordinar esfuerzos interinstitucionales para prevenir y reducir el robo de motocicletas.",
        "zona_trabajo": "Santa Teresa",
        "actores": "Tránsito/Migración",
        "indicador_actividad": "Operativos",
        "consideraciones": "No aplica",
        "periodicidad": "2 por semana",
        "meta_cuantitativa": 20,
        "responsable": "Dirección Regional",
        "efecto_esperado": (
            "1- Disminución de la tasa de robo de motocicletas. "
            "2- Incremento en la detención y judicialización de los responsables. "
            "3- Desarticulación de bandas dedicadas a este delito."
        ),
    },
    {
        "fila": 2,
        "indole": "Operativo",
        "actividad_estrategica": "Intensificar los operativos de investigación conjuntos con OIJ para captura de responsables y recuperación de motocicletas robadas.",
        "zona_trabajo": "Santa Teresa",
        "actores": "OIJ",
        "indicador_actividad": "Operativos",
        "consideraciones": "No aplica",
        "periodicidad": "1 por quincena",
        "meta_cuantitativa": 10,
        "responsable": "Dirección Regional",
        "efecto_esperado": (
            "1- Desarticulación de estructuras criminales. "
            "2- Creación de un fuerte efecto disuasorio. "
            "3- Disminución estadística del delito."
        ),
    },
    {
        "fila": 3,
        "indole": "Operativo",
        "actividad_estrategica": "Implementar sistema de registro y fiscalización georreferenciado de talleres y chatarreras para prevenir venta de partes y motocicletas de procedencia ilícita.",
        "zona_trabajo": "Santa Teresa",
        "actores": "OIJ",
        "indicador_actividad": "Informe realizado",
        "consideraciones": "Destacar la georreferenciación actualizada de los lugares de interés policial.",
        "periodicidad": "1 bimensual actualizado y georreferenciado",
        "meta_cuantitativa": 2,
        "responsable": "Dirección Regional",
        "efecto_esperado": (
            "1- Reducción sostenida del delito de robo de motocicletas. "
            "2- Fortalecimiento de la capacidad de control del Estado."
        ),
    },
    {
        "fila": 4,
        "indole": "Operativo",
        "actividad_estrategica": "Identificar, geolocalizar y categorizar puntos de búnkers y casas de venta de droga para optimizar operativos y desarticulación de redes.",
        "zona_trabajo": "Santa Teresa",
        "actores": "OIJ",
        "indicador_actividad": "Informe realizado",
        "consideraciones": "Destacar la georreferenciación actualizada de los lugares de interés policial.",
        "periodicidad": "1 bimensual actualizado y georreferenciado",
        "meta_cuantitativa": 2,
        "responsable": "Dirección Regional",
        "efecto_esperado": (
            "1- Creación de un Mapa Dinámico de venta de droga. "
            "2- Optimización de recursos policiales. "
            "3- Análisis predictivo. "
            "4- Aumento en la efectividad de acciones policiales y allanamientos."
        ),
    },
    {
        "fila": 5,
        "indole": "Operativo",
        "actividad_estrategica": "Plan de intervención interinstitucional en bares para prevención de delitos, narcomenudeo y actos de violencia.",
        "zona_trabajo": "Santa Teresa",
        "actores": "FP/Turística/Tránsito/OIJ",
        "indicador_actividad": "Operativos",
        "consideraciones": "No aplica",
        "periodicidad": "1 bimensual",
        "meta_cuantitativa": 2,
        "responsable": "Dirección Regional",
        "efecto_esperado": (
            "1- Desplazamiento de la actividad criminal. "
            "2- Reducción de la violencia y riñas. "
            "3- Efecto disuasorio. "
            "4- Prevención del narcomenudeo."
        ),
    },
]

def get_conn():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA foreign_keys=ON;")
    return conn

def _col_exists(cur, table, col):
    cur.execute(f"PRAGMA table_info({table});")
    cols = [r[1] for r in cur.fetchall()]
    return col in cols

def init_db():
    conn = get_conn()
    cur = conn.cursor()
    # Tabla metas con columnas extendidas
    cur.execute("""
        CREATE TABLE IF NOT EXISTS metas (
            fila INTEGER PRIMARY KEY,
            actividad TEXT NOT NULL,      -- alias de actividad_estrategica
            meta_total INTEGER NOT NULL,  -- alias de meta_cuantitativa
            indole TEXT,
            zona_trabajo TEXT,
            actores TEXT,
            indicador_actividad TEXT,
            consideraciones TEXT,
            periodicidad TEXT,
            responsable TEXT,
            efecto_esperado TEXT
        );
    """)
    # Migraciones suaves (si existía tabla vieja)
    needed = [
        ("indole", "TEXT"),
        ("zona_trabajo", "TEXT"),
        ("actores", "TEXT"),
        ("indicador_actividad", "TEXT"),
        ("consideraciones", "TEXT"),
        ("periodicidad", "TEXT"),
        ("responsable", "TEXT"),
        ("efecto_esperado", "TEXT"),
    ]
    for col, typ in needed:
        if not _col_exists(cur, "metas", col):
            cur.execute(f"ALTER TABLE metas ADD COLUMN {col} {typ};")

    cur.execute("""
        CREATE TABLE IF NOT EXISTS movimientos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fila INTEGER NOT NULL,
            fecha TEXT NOT NULL,            -- DD-MM-YYYY
            cantidad INTEGER NOT NULL CHECK(cantidad >= 0),
            nota TEXT,
            delta INTEGER NOT NULL,         -- con signo
            FOREIGN KEY(fila) REFERENCES metas(fila)
        );
    """)
    conn.commit()

    # Seed si está vacío
    cur.execute("SELECT COUNT(*) FROM metas;")
    if cur.fetchone()[0] == 0:
        cur.executemany(
            """
            INSERT INTO metas
            (fila, actividad, meta_total, indole, zona_trabajo, actores, indicador_actividad,
             consideraciones, periodicidad, responsable, efecto_esperado)
            VALUES
            (:fila, :actividad, :meta_total, :indole, :zona_trabajo, :actores, :indicador_actividad,
             :consideraciones, :periodicidad, :responsable, :efecto_esperado)
            """,
            [
                {
                    "fila": it["fila"],
                    "actividad": it["actividad_estrategica"],
                    "meta_total": int(it["meta_cuantitativa"] or 0),
                    "indole": it.get("indole", ""),
                    "zona_trabajo": it.get("zona_trabajo", ""),
                    "actores": it.get("actores", ""),
                    "indicador_actividad": it.get("indicador_actividad", ""),
                    "consideraciones": it.get("consideraciones", ""),
                    "periodicidad": it.get("periodicidad", ""),
                    "responsable": it.get("responsable", ""),
                    "efecto_esperado": it.get("efecto_esperado", ""),
                }
                for it in PLAN_BASE
            ]
        )
        conn.commit()
    conn.close()

init_db()

# =========================
# 2) CONSULTAS / ACCIONES DB
# =========================
def obtener_metas_df() -> pd.DataFrame:
    conn = get_conn()
    df = pd.read_sql_query("""
        SELECT fila, actividad, meta_total,
               indole, zona_trabajo, actores, indicador_actividad,
               consideraciones, periodicidad, responsable, efecto_esperado
        FROM metas
        ORDER BY fila;
    """, conn)
    conn.close()
    return df

def suma_delta_por_fila(fila: int) -> int:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT COALESCE(SUM(delta), 0) FROM movimientos WHERE fila=?;", (fila,))
    total = cur.fetchone()[0] or 0
    conn.close()
    return int(total)

def obtener_historial(fila: int) -> List[Dict[str, Any]]:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, fecha, cantidad, nota, delta
        FROM movimientos
        WHERE fila=?
        ORDER BY id ASC;
    """, (fila,))
    rows = cur.fetchall()
    conn.close()
    return [
        {"id": r[0], "fecha": r[1], "cantidad": int(r[2]), "nota": r[3] or "", "delta": int(r[4])}
        for r in rows
    ]

def meta_total_de_fila(fila: int) -> int:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT meta_total FROM metas WHERE fila=?;", (fila,))
    row = cur.fetchone()
    conn.close()
    return int(row[0]) if row else 0

def insertar_movimiento(fila: int, mov: int, nota: str) -> bool:
    meta_total = meta_total_de_fila(fila)
    avance_actual = suma_delta_por_fila(fila)
    nuevo_avance = max(0, min(meta_total, avance_actual + int(mov)))
    delta_real = int(nuevo_avance - avance_actual)
    if delta_real == 0 and not (nota or "").strip():
        return False
    fecha = datetime.now().strftime("%d-%m-%Y")
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO movimientos (fila, fecha, cantidad, nota, delta)
        VALUES (?, ?, ?, ?, ?);
    """, (fila, fecha, abs(delta_real), (nota or "").strip(), delta_real))
    conn.commit()
    conn.close()
    return True

def actualizar_movimiento(id_mov: int, fila: int, nueva_cant: int, nueva_nota: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT delta FROM movimientos WHERE id=?;", (id_mov,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return
    old_delta = int(row[0])
    sign = 1 if old_delta >= 0 else -1
    cur.execute("SELECT COALESCE(SUM(delta),0) FROM movimientos WHERE fila=? AND id<>?;", (fila, id_mov))
    avance_sin = int(cur.fetchone()[0] or 0)
    meta_total = meta_total_de_fila(fila)
    nuevo_delta_deseado = sign * int(nueva_cant)
    min_allowed = -avance_sin
    max_allowed = meta_total - avance_sin
    nuevo_delta = max(min_allowed, min(max_allowed, nuevo_delta_deseado))
    nueva_cant_recortada = abs(int(nuevo_delta))
    cur.execute("""
        UPDATE movimientos
        SET cantidad = ?, nota = ?, delta = ?
        WHERE id = ?;
    """, (nueva_cant_recortada, (nueva_nota or "").strip(), nuevo_delta, id_mov))
    conn.commit()
    conn.close()

def eliminar_movimiento(id_mov: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM movimientos WHERE id=?;", (id_mov,))
    conn.commit()
    conn.close()

def obtener_resumen_df() -> pd.DataFrame:
    metas = obtener_metas_df()
    conn = get_conn()
    avances = pd.read_sql_query("""
        SELECT fila, COALESCE(SUM(delta),0) AS avance
        FROM movimientos
        GROUP BY fila;
    """, conn)
    conn.close()
    df = metas.merge(avances, on="fila", how="left").fillna({"avance": 0})
    df["avance"] = df["avance"].astype(int)
    df["limite_restante"] = df["meta_total"] - df["avance"]
    df["porcentaje_val"] = (df["avance"] / df["meta_total"].replace(0, pd.NA) * 100).astype(float).round(1)
    df["porcentaje_val"] = df["porcentaje_val"].fillna(0.0)
    df["porcentaje"] = df["porcentaje_val"].map(lambda x: f"{x:.1f}%")
    df["estado"] = df.apply(
        lambda r: "Completa" if r["porcentaje_val"] >= 100 else ("En curso" if r["avance"] > 0 else "Pendiente"),
        axis=1
    )
    return df.sort_values("fila").reset_index(drop=True)

# ==================================
# 3) ESTADO DE UI (inputs por actividad)
# ==================================
if "reset_flags" not in st.session_state:
    st.session_state["reset_flags"] = {}

def set_reset_flag(fila: int, val: bool):
    st.session_state["reset_flags"][fila] = val

def get_reset_flag(fila: int) -> bool:
    return st.session_state["reset_flags"].get(fila, False)

def ensure_ui_keys_for_fila(fila: int):
    st.session_state.setdefault(f"mov_val_{fila}", 0)
    st.session_state.setdefault(f"nota_inline_{fila}", "")

# =========================
# 4) UI PRINCIPAL POR FILA
# =========================
df_base = obtener_metas_df()

for _, r in df_base.iterrows():
    f = int(r["fila"])
    ensure_ui_keys_for_fila(f)

    if get_reset_flag(f):
        st.session_state[f"mov_val_{f}"] = 0
        st.session_state[f"nota_inline_{f}"] = ""
        set_reset_flag(f, False)

    meta_total = int(r["meta_total"])
    avance = int(suma_delta_por_fila(f))
    restante = meta_total - avance

    colA, colB = st.columns([2.2, 1])
    with colA:
        st.markdown(f"**{r['actividad']}**  \nMeta original: **{meta_total}**")
        st.caption(f"Índole: {r['indole']} • Periodicidad: {r['periodicidad']} • Indicador: {r['indicador_actividad']}")
    with colB:
        st.metric("Límite restante", restante)

    c1, c2, c3 = st.columns([1.1, 2.2, 1])
    with c1:
        st.number_input(
            "Movimiento",
            key=f"mov_val_{f}",
            step=1, format="%d",
            min_value=-meta_total,
            max_value= meta_total,
            help="− resta (avanza), + suma (devuelve). Empieza en 0."
        )
    with c2:
        st.text_input(
            "Nota del movimiento (opcional)",
            key=f"nota_inline_{f}",
            placeholder="Breve descripción…"
        )
    with c3:
        if st.button("Guardar movimiento", key=f"guardar_{f}"):
            mov = int(st.session_state[f"mov_val_{f}"])
            nota_mov = (st.session_state[f"nota_inline_{f}"] or "").strip()
            inserted = insertar_movimiento(f, mov, nota_mov)
            set_reset_flag(f, True)
            st.rerun()

    st.divider()

# =========================
# 5) TABLA RESUMEN
# =========================
df = obtener_resumen_df()
st.dataframe(
    df[["actividad", "meta_total", "avance", "limite_restante", "porcentaje", "estado"]],
    use_container_width=True
)

# =========================
# 6) BURBUJAS: VER/EDITAR/ELIMINAR HISTORIAL
# =========================
st.markdown("#### Resumen interactivo (clic en el **avance** para ver/editar historial)")

for _, row in df.iterrows():
    f = int(row["fila"])
    c1, c2, c3, c4, c5, c6 = st.columns([4, 1.1, 1.1, 1.1, 1.2, 1.8])
    with c1:
        st.markdown(f"**{row['actividad']}**")
    with c2:
        st.caption("meta")
        st.write(int(row["meta_total"]))
    with c3:
        st.caption("límite restante")
        st.write(int(row["limite_restante"]))
    with c4:
        st.caption("avance")
        with st.popover(f"{int(row['avance'])}"):
            st.markdown(f"**Historial — {row['actividad']}**")
            hist = obtener_historial(f)
            st.caption(f"Movimientos registrados: {len(hist)}")
            if not hist:
                st.caption("Sin movimientos registrados aún.")
            else:
                df_hist_view = pd.DataFrame([
                    {"Fecha": i["fecha"], "Cantidad": i["cantidad"], "Nota": i.get("nota", "")}
                    for i in hist
                ])
                st.table(df_hist_view)

                st.markdown("**Editar / eliminar**")
                for item in hist:
                    id_mov = int(item["id"])
                    ec1, ec2, ec3, ec4 = st.columns([1, 1, 3, 1.2])
                    with ec1:
                        st.text_input("Fecha", value=item["fecha"], key=f"edit_fecha_{f}_{id_mov}", disabled=True)
                    with ec2:
                        nueva_cant = st.number_input(
                            "Cantidad", min_value=0, step=1,
                            value=int(item["cantidad"]),
                            key=f"edit_cant_{f}_{id_mov}"
                        )
                    with ec3:
                        nueva_nota = st.text_input(
                            "Nota", value=item.get("nota",""),
                            key=f"edit_nota_{f}_{id_mov}"
                        )
                    with ec4:
                        if st.button("💾 Guardar", key=f"save_edit_{f}_{id_mov}"):
                            actualizar_movimiento(id_mov, f, int(nueva_cant), nueva_nota)
                            st.rerun()
                        if st.button("🗑️ Eliminar", key=f"del_{f}_{id_mov}"):
                            eliminar_movimiento(id_mov)
                            st.rerun()

    with c5:
        st.caption("porcentaje")
        st.write(row["porcentaje"])
    with c6:
        st.caption("estado")
        st.write(row["estado"])
    st.divider()

# =========================
# 7) MÉTRICA GLOBAL
# =========================
meta_total_sum = int(df["meta_total"].sum())
avance_total = int(df["avance"].sum())
pct_total = (avance_total / meta_total_sum) * 100 if meta_total_sum else 0
st.metric("Avance total (todas las metas)", f"{pct_total:.1f}%")

# =========================
# 8) DESCARGAR EXCEL (multi-hoja, sin 'delta', con estilos)
# =========================
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

buffer = BytesIO()

# --- Hoja RESUMEN (igual a tu tabla + contexto) ---
df_resumen = df[[
    "fila", "actividad", "meta_total", "avance", "limite_restante", "porcentaje", "estado"
]].copy()

# Agregar columnas de contexto
ctx = obtener_metas_df().set_index("fila")
for col in ["indole", "zona_trabajo", "actores", "indicador_actividad",
            "consideraciones", "periodicidad", "responsable", "efecto_esperado"]:
    df_resumen[col] = df_resumen["fila"].map(ctx[col])

# --- Hoja HISTORIAL (1 fila por movimiento, SIN 'delta') ---
hist_rows = []
for _, r in df.iterrows():
    f = int(r["fila"])
    actividad = r["actividad"]
    for m in obtener_historial(f):
        hist_rows.append({
            "fila": f,
            "actividad": actividad,
            "fecha": m.get("fecha", ""),
            "cantidad": int(m.get("cantidad", 0)),
            "nota": m.get("nota", ""),
        })
df_hist = pd.DataFrame(hist_rows)

# --- Hoja RESPALDO (solo notas no vacías) ---
if not df_hist.empty:
    df_respaldo = df_hist[df_hist["nota"].astype(str).str.strip() != ""].loc[:, ["fila", "actividad", "fecha", "nota"]].copy()
else:
    df_respaldo = pd.DataFrame(columns=["fila", "actividad", "fecha", "nota"])

def estilizar_hoja(ws, hex_tab):
    # Color de pestaña
    ws.sheet_properties.tabColor = hex_tab
    # Estilos de encabezado
    header_fill = PatternFill("solid", fgColor="1E88E5")  # azul
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = align_center
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = max(12, min(60, len(str(c.value)) + 6))
    ws.freeze_panes = "A2"

with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
    df_resumen.to_excel(writer, index=False, sheet_name="Resumen")
    if not df_hist.empty:
        df_hist.to_excel(writer, index=False, sheet_name="Historial")
    if not df_respaldo.empty:
        df_respaldo.to_excel(writer, index=False, sheet_name="Respaldo (notas)")

    # Aplicar colores a pestañas + encabezados
    if "Resumen" in writer.sheets:
        estilizar_hoja(writer.sheets["Resumen"], "1E88E5")      # azul
    if "Historial" in writer.sheets:
        estilizar_hoja(writer.sheets["Historial"], "E53935")    # rojo
    if "Respaldo (notas)" in writer.sheets:
        estilizar_hoja(writer.sheets["Respaldo (notas)"], "43A047")  # verde

buffer.seek(0)
st.download_button(
    "📥 Descargar desglose en Excel",
    buffer,
    file_name="avance_por_meta_movimientos.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

# =========================
# 9) 📊 Visualizaciones por meta (ocultas hasta seleccionar)
# =========================
st.markdown("### 📊 Visualizaciones por meta")

import matplotlib.pyplot as plt
import numpy as np

# Colores vivos
BLUE = "#1E88E5"   # azul intenso
RED  = "#E53935"   # rojo intenso

def _prep_fig():
    fig, ax = plt.subplots(figsize=(8, 4.5), facecolor="black")
    ax.set_facecolor("black")
    ax.tick_params(colors="white")
    ax.spines["bottom"].set_color("white")
    ax.spines["left"].set_color("white")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].setVisible = False
    ax.grid(axis="y", alpha=0.15, color="white")
    return fig, ax

_df_opts = df[["fila", "actividad", "meta_total", "avance", "limite_restante", "porcentaje_val"]].copy()
_df_opts["op"] = _df_opts["fila"].astype(str) + " — " + _df_opts["actividad"]
placeholder = "— Selecciona una meta —"
options = [placeholder] + _df_opts["op"].tolist()
sel = st.selectbox("Elegí la meta a visualizar", options, index=0, key="sel_meta_uno")

if sel == placeholder:
    st.info("Seleccioná una meta para mostrar el gráfico.")
else:
    fila_sel = int(_df_opts.loc[_df_opts["op"] == sel, "fila"].iloc[0])
    row_sel = df.loc[df["fila"] == fila_sel].iloc[0]

    meta = int(row_sel["meta_total"])
    avance = int(row_sel["avance"])
    restante = max(0, meta - avance)
    pct = float(row_sel["porcentaje_val"])

    tipo = st.radio("Tipo de gráfico", ["Barras", "Circular"], index=0, horizontal=True, key="tipo_uno_por_uno")

    if tipo == "Barras":
        fig, ax = _prep_fig()
        vals = [avance, restante]
        labels = ["Avance", "Restante"]
        x = np.arange(len(labels))
        width = 0.6
        ax.bar(x + 0.03, vals, width=width, color="black", alpha=0.35, zorder=0)  # sombra
        bars = ax.bar(x, vals, width=width, color=[BLUE, RED], alpha=0.95, edgecolor="white", linewidth=1.2, zorder=1)
        y_max = max(meta, max(vals), 1)
        ax.set_ylim(0, y_max * 1.15)
        ax.set_xticks(x)
        ax.set_xticklabels(labels, color="white")
        ax.set_ylabel("Cantidad", color="white")
        ax.set_title(f"{row_sel['actividad']} — Meta {meta}  |  Avance total: {pct:.1f}%", color="white")
        for b, val in zip(bars, vals):
            perc = (val / meta * 100) if meta else 0.0
            ax.text(b.get_x() + b.get_width()/2, b.get_height() + (y_max * 0.03),
                    f"{val}  ({perc:.1f}%)", ha="center", va="bottom", color="white", fontsize=10)
        st.pyplot(fig, clear_figure=True)

    elif tipo == "Circular":
        fig, ax = _prep_fig()
        datos = [max(avance, 0), max(restante, 0)]
        etiquetas = ["Avance", "Restante"]
        if sum(datos) == 0:
            datos, etiquetas = [1], ["Sin datos"]
        def autopct_fmt(p): return f"{p:.1f}%"
        wedges, texts, autotexts = ax.pie(
            datos, labels=etiquetas, autopct=autopct_fmt, startangle=90,
            colors=[BLUE, RED], shadow=True, wedgeprops=dict(edgecolor="white", linewidth=1.2)
        )
        for t in texts + autotexts:
            t.set_color("white")
        ax.axis("equal")
        ax.set_title(f"{row_sel['actividad']} — Meta {meta}  |  Avance total: {pct:.1f}%", color="white")
        st.pyplot(fig, clear_figure=True)



